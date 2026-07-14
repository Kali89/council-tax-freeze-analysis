"""
Parses MHCLG's Core Spending Power (CSP) table into a per-2025-LA,
per-financial-year settlement-funding series - the control variable for
the Phase 6 / Variant 3 regression (did rate-setting already offset the
freeze's redistributive effect?).

Source: the FINAL 2025-26 settlement's own CSP table
(`CSP_information_table_LGFS_2025-26.xlsx`) is a maintained "live table",
like Band D's - one workbook, one sheet per year, back to 2015-16, not 11
separate per-vintage releases. Confirmed directly: every sheet has a
consistent `ons_code` column and a same-year CSP total column named
`csp_{year_start}` (e.g. sheet "2018-19" has its own year's total in
column `csp_2018`), even though the SUPPORTING grant-component columns
around it change name and composition every year (Revenue Support Grant
rolled into Settlement Funding Assessment, New Homes Bonus phased out,
Social Care Grant introduced, etc. - the underlying policy churn DATA.md
already flagged). **Pre-2015-16 is not in this file and is out of scope
by design** - "Core Spending Power" as a concept didn't exist before
2015-16 (earlier years used Formula Grant, then early Settlement Funding
Assessment); see DATA.md "Variant 3" for why the regression is reported
on 2015-16 to 2025-26 rather than blocked on reconstructing an
incompatible pre-2015 series from the Web Archive.

**Tiering, and why this parser needs CTSOP dwelling counts as an input.**
Each sheet has FOUR kinds of row, confirmed by checking every code prefix
directly: E06/E07/E08/E09 (296 district/unitary/borough rows - the SAME
2025-vintage set as everywhere else in this pipeline, though EARLIER
sheets still use predecessor codes for LAs that later reorganised,
resolved via the Phase 1 crosswalk exactly like Band D/CTSOP), E10 (21-27
shire COUNTY rows), a single row for the Greater London Authority (code
"-", not a real GSS code), and (from 2019-20) a single row for Greater
Manchester Combined Authority (E47000001). E31 (fire authority) rows are
DELIBERATELY EXCLUDED - fire is a genuinely separate precepting/funding
stream, not part of a council's own rate-setting decision.

The upper-tier rows are NOT a rounding error: checked directly,
Cambridgeshire's own county row (GBP595.0m, 2025-26) is **7.1x** its five
districts' own CSP rows combined (GBP83.5m), and the GLA's own row
(GBP3,212.2m) is **31.7%** of the 33 London boroughs' own rows combined
(GBP10,139.0m) - the same "own precept is a small slice of the area
total" structure Phase 4 found and quantified for Band D
(`engine.build.compute_shared_tier_exposure`). Using only the
district/unitary/borough row would represent well under 20% of true
funding for most of England. So each upper-tier row is APPORTIONED to
its constituent 2025-vintage LAs by dwelling share (CTSOP `all_properties`
- the same weighting logic already established for reallocating shared
tiers elsewhere in this pipeline), using
`boundaries.precepting_groups.PRECEPTING_GROUP` to know which LAs share
which upper tier.

**A known, quantified gap in that apportionment - not silently fixed.**
`PRECEPTING_GROUP` reflects the CURRENT (December 2024) tiering
structure only. For the handful of counties that unitarised DURING the
2015-16 to 2025-26 window (Northamptonshire 2021, Dorset 2019, Cumbria /
North Yorkshire / Somerset 2023), their predecessor districts resolve
(via the Phase 1 crosswalk) to a 2025 target that is NOW standalone
unitary in `PRECEPTING_GROUP` - so for years BEFORE that LA's own reorg,
this parser does not apportion that year's real county CSP row to it,
even though it genuinely was two-tier at the time. `coverage` reports
exactly which (predecessor code, year) cells this affects, and DATA.md
quantifies the resulting understatement, following the same
quantify-before-deciding discipline as the Suffolk/Somerset 2019 CTSOP
gap in Phase 4 - not assumed to be immaterial, checked.
"""

from __future__ import annotations

from dataclasses import dataclass

import openpyxl
import pandas as pd

from council_tax_freeze.boundaries.crosswalk import AmbiguousBoundaryChange, UnmappedLocalAuthorityCode, resolve
from council_tax_freeze.boundaries.lad_2025 import LAD_2025_CODES
from council_tax_freeze.boundaries.precepting_groups import PRECEPTING_GROUP

OWN_TIER_CODE_RE = r"^E0[6-9]\d{6}$"
COUNTY_CODE_RE = r"^E10\d{6}$"
COMBINED_AUTHORITY_CODE_RE = r"^E47\d{6}$"
GLA_ROW_CODE = "-"

# Barnsley/Sheffield's 2025 boundary change (the Oughtibridge Mill
# transfer, reorg_events.py BARNSLEY_SHEFFIELD_2025) is dated 2025-04-01,
# and crosswalk.resolve() treats a query AT (not just before) an event's
# own effective_date as not-yet-applicable (a pre-existing Phase 1
# convention, not something to patch here) - so `resolve(..., "2025-04-01",
# ...)` returns Barnsley's OLD code unchanged rather than applying the
# split. Band D never hits this because its source data already applies
# the new codes retroactively; this settlement source still lists Barnsley/
# Sheffield under their OLD codes even in the 2025-26 sheet, so this
# parser needs the same one-off substitution engine/build.py already uses
# for the CTSOP join, for the same reason.
BARNSLEY_SHEFFIELD_2025_CODE_SUBSTITUTION = {"E08000016": "E08000038", "E08000019": "E08000039"}

# Predecessor codes of the 5 counties that unitarised DURING the 2015-16 to
# 2025-26 window (Dorset 2019, Northamptonshire 2021, Cumbria/North
# Yorkshire/Somerset 2023) - genuinely two-tier shire districts, per
# reorg_events.py, as opposed to predecessors that were ALREADY unitary
# before their own merge (Bournemouth E06000028, Poole E06000029 - both
# already standalone, correctly excluded here since there is no county
# apportionment gap for them to begin with). Used only to give an exact,
# not heuristic, count of the known transitional-county apportionment gap
# - see module docstring.
TRANSITIONAL_COUNTY_PREDECESSOR_CODES = frozenset(
    {
        # Dorset, pre-2019-04-01
        "E07000048", "E07000049", "E07000050", "E07000051", "E07000052", "E07000053",
        # Somerset, pre-2023-04-01 (including the 2019-created intermediate Somerset West and Taunton)
        "E07000187", "E07000188", "E07000189", "E07000190", "E07000191", "E07000246",
        # Northamptonshire (both 2021 successors), pre-2021-04-01
        "E07000150", "E07000151", "E07000152", "E07000153", "E07000154", "E07000155", "E07000156",
        # Cumbria (both 2023 successors), pre-2023-04-01
        "E07000026", "E07000027", "E07000028", "E07000029", "E07000030", "E07000031",
        # North Yorkshire, pre-2023-04-01
        "E07000163", "E07000164", "E07000165", "E07000166", "E07000167", "E07000168", "E07000169",
    }
)

# Sheet names in the workbook, oldest first - "2015-16" is the first year
# Core Spending Power exists as a concept.
SHEET_NAMES = [f"{y}-{str(y + 1)[2:]}" for y in range(2015, 2026)]


@dataclass
class SettlementResult:
    la_year: pd.DataFrame  # ons_code, financial_year, csp_own, csp_upper_tier_apportioned, csp_total (GBP)
    coverage: pd.DataFrame  # financial_year, n_la, n_unapportioned_upper_tier_gap (see module docstring)
    unresolved: pd.DataFrame  # diagnostic - predecessor rows the crosswalk couldn't resolve


def _read_sheet_rows(path, sheet: str) -> list[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    wb.close()
    return rows


def _sheet_columns(rows: list[tuple]) -> dict[str, int]:
    """Row index 3 (0-based) holds the machine-readable column codes
    (`ons_code`, `authority`, `csp_{year}`, ...) - consistent across every
    sheet, confirmed directly, even though the human-readable label row
    above it (row 4) and the surrounding grant-component columns change
    every year."""
    codes_row = rows[3]
    return {c: i for i, c in enumerate(codes_row) if c}


def _year_start(financial_year: str) -> int:
    return int(financial_year.split("-")[0])


def _parse_sheet(path, financial_year: str) -> pd.DataFrame:
    rows = _read_sheet_rows(path, financial_year)
    col = _sheet_columns(rows)
    csp_col = f"csp_{_year_start(financial_year)}"
    oc, auth, csp = col["ons_code"], col["authority"], col[csp_col]

    out = []
    for r in rows[8:]:  # data starts after the "England" total row (row index 8, 0-based)
        code, name, value = r[oc], r[auth], r[csp]
        if code in (None, ""):
            continue
        out.append({"code": str(code), "name": name, "csp": float(value) if value is not None else None})
    return pd.DataFrame(out)


def build_settlement(path, ctsop_la_year: pd.DataFrame) -> SettlementResult:
    dwellings = ctsop_la_year.set_index(["ons_code", "financial_year"])["all_properties"]

    own_tier_rows = []
    unresolved = []
    unapportioned_gap = []

    for fy in SHEET_NAMES:
        sheet_df = _parse_sheet(path, fy)
        as_of = f"{_year_start(fy)}-04-01"

        own_tier = sheet_df[sheet_df["code"].str.match(OWN_TIER_CODE_RE)]
        county = sheet_df[sheet_df["code"].str.match(COUNTY_CODE_RE)].set_index("name")["csp"]
        combined_authority = sheet_df[sheet_df["code"].str.match(COMBINED_AUTHORITY_CODE_RE)].set_index("name")["csp"]
        gla_rows = sheet_df[sheet_df["code"] == GLA_ROW_CODE]
        gla_csp = gla_rows["csp"].iloc[0] if len(gla_rows) else None

        upper_tier_by_group = dict(county)
        if len(combined_authority):
            upper_tier_by_group["Greater Manchester"] = upper_tier_by_group.get("Greater Manchester", 0) + combined_authority.get(
                "Greater Manchester Combined Authority", 0
            )
        if gla_csp is not None:
            upper_tier_by_group["Greater London"] = upper_tier_by_group.get("Greater London", 0) + gla_csp

        # Resolve each own-tier row's predecessor code to its 2025 target(s),
        # exactly the same crosswalk pattern as Band D/CTSOP - EXCEPT this
        # data source, unlike Band D, uses Barnsley's OLD code (E08000016)
        # even in recent-year sheets rather than the new one applied
        # retroactively, so resolving it here (unlike in engine/build.py)
        # actually walks through the 2025 Barnsley/Sheffield SPLIT event and
        # needs a year_totals figure for its fixed_transfer apportionment
        # (the 12-dwelling Oughtibridge Mill transfer) - same stub value
        # already established in tests/test_boundaries.py; the exact number
        # is immaterial here since CSP is measured in GBP millions and a
        # 12-dwelling transfer changes the resulting weight by a
        # vanishingly small amount regardless of what "total" it's divided
        # by.
        resolved_rows = []  # (target_code, csp_own_share, precepting_group_of_target)
        for _, r in own_tier.iterrows():
            code, name, csp_value = r["code"], r["name"], r["csp"]
            if fy == "2025-26" and code in BARNSLEY_SHEFFIELD_2025_CODE_SUBSTITUTION:
                resolved_rows.append((BARNSLEY_SHEFFIELD_2025_CODE_SUBSTITUTION[code], csp_value or 0, PRECEPTING_GROUP.get(BARNSLEY_SHEFFIELD_2025_CODE_SUBSTITUTION[code])))
                continue
            try:
                resolved = resolve(code, name, as_of, year_totals={"Barnsley": 1_000_000})
            except (AmbiguousBoundaryChange, UnmappedLocalAuthorityCode) as e:
                unresolved.append({"code": code, "authority": name, "financial_year": fy, "reason": str(e)})
                continue
            for target_code, _target_name, weight in resolved:
                if target_code not in LAD_2025_CODES:
                    unresolved.append({"code": code, "authority": name, "financial_year": fy, "reason": f"resolved to non-2025 code {target_code}"})
                    continue
                group = PRECEPTING_GROUP.get(target_code)
                own_share = (csp_value or 0) * weight
                resolved_rows.append((target_code, own_share, group))
                if code in TRANSITIONAL_COUNTY_PREDECESSOR_CODES:
                    # genuinely two-tier at this date (per reorg_events.py), but its
                    # 2025 target is standalone in PRECEPTING_GROUP - the known,
                    # quantified apportionment gap. See module docstring.
                    unapportioned_gap.append({"code": code, "authority": name, "financial_year": fy, "target_code": target_code})

        target_own = {}
        target_group = {}
        for target_code, own_share, group in resolved_rows:
            target_own[target_code] = target_own.get(target_code, 0) + own_share
            target_group[target_code] = group

        # Apportion each upper-tier row across its CURRENT (2025) precepting
        # group's members by dwelling share - see module docstring for the
        # known gap this creates for transitional counties in years before
        # their own reorg.
        group_members: dict[str, list[str]] = {}
        for code, group in target_group.items():
            if group and not group.startswith("__standalone__"):
                group_members.setdefault(group, []).append(code)

        apportioned = dict.fromkeys(target_own, 0.0)
        for group, members in group_members.items():
            upper_csp = upper_tier_by_group.get(group)
            if not upper_csp:
                continue
            weights = {}
            for code in members:
                w = dwellings.get((code, fy))
                weights[code] = w if w is not None else 0
            total_w = sum(weights.values())
            if total_w <= 0:
                continue
            for code in members:
                apportioned[code] += upper_csp * (weights[code] / total_w)

        for code in target_own:
            own_tier_rows.append(
                {
                    "ons_code": code,
                    "financial_year": fy,
                    "csp_own": target_own[code],
                    "csp_upper_tier_apportioned": apportioned.get(code, 0.0),
                }
            )

    la_year = pd.DataFrame(own_tier_rows)
    la_year["csp_total"] = la_year["csp_own"] + la_year["csp_upper_tier_apportioned"]

    coverage = (
        la_year.groupby("financial_year", as_index=False)
        .agg(n_la=("ons_code", "nunique"))
        .merge(
            pd.DataFrame(unapportioned_gap).groupby("financial_year", as_index=False).size().rename(columns={"size": "n_unapportioned_upper_tier_gap"})
            if unapportioned_gap
            else pd.DataFrame({"financial_year": SHEET_NAMES, "n_unapportioned_upper_tier_gap": 0}),
            on="financial_year",
            how="left",
        )
    )
    coverage["n_unapportioned_upper_tier_gap"] = coverage["n_unapportioned_upper_tier_gap"].fillna(0).astype(int)

    return SettlementResult(la_year=la_year, coverage=coverage, unresolved=pd.DataFrame(unresolved))
